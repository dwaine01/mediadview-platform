# ruff: noqa: E701,E702,E741,E731,F811,W293,W605,I001
"""
MediAd View — Reports export helpers (Fase 5 · Sprint 1 · Etapa C4).

Every function is a PURE bytes-generator. Routes marshal the payload
into a `fastapi.Response` with the right content-type/disposition.

Formats supported: CSV, XLSX (openpyxl), PDF (reportlab).
Rows may contain datetimes, floats, ints, strings, None.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

CT_CSV  = "text/csv; charset=utf-8"
CT_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CT_PDF  = "application/pdf"


def _stringify(v: Any) -> str:
    if v is None: return ""
    if isinstance(v, datetime):
        return v.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    if isinstance(v, bool):
        return "true" if v else "false"
    return str(v)


def _fmt_money(cents: Any, currency: str = "USD") -> str:
    if cents is None: return ""
    try:
        n = int(cents)
    except Exception:
        return str(cents)
    return f"{n/100:,.2f} {(currency or 'USD').upper()}"


# ═══════════════════════════════════════════════════════════════════
# CSV
# ═══════════════════════════════════════════════════════════════════
def to_csv(rows: list[dict], columns: list[str]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow(columns)
    for r in rows:
        w.writerow([_stringify(r.get(c)) for c in columns])
    return buf.getvalue().encode("utf-8")


# ═══════════════════════════════════════════════════════════════════
# XLSX
# ═══════════════════════════════════════════════════════════════════
def to_xlsx(sheets: list[dict]) -> bytes:
    """`sheets` is a list of dicts: {name, rows, columns, [column_widths]}."""
    wb = Workbook()
    # Drop default sheet if we create real ones
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="E2E8F0", size=10)
    header_align = Alignment(horizontal="left", vertical="center")

    for sh in sheets:
        name = (sh.get("name") or "Sheet")[:31]
        ws = wb.create_sheet(title=name)
        cols = sh["columns"]
        for col_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        for row_idx, r in enumerate(sh.get("rows", []), start=2):
            for col_idx, col in enumerate(cols, start=1):
                v = r.get(col)
                if isinstance(v, datetime):
                    v = v.astimezone(timezone.utc).replace(tzinfo=None)
                ws.cell(row=row_idx, column=col_idx, value=v)
        # Auto widths (best-effort)
        widths = sh.get("column_widths") or {}
        for col_idx, col in enumerate(cols, start=1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = widths.get(col, max(12, min(48, len(col) + 4)))
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════
# PDF (Landscape LETTER for readability of long tables)
# ═══════════════════════════════════════════════════════════════════
def to_pdf(*, title: str, subtitle: str, sections: list[dict]) -> bytes:
    """`sections` is a list of dicts:
        {"heading": str, "rows": [dict], "columns": [str],
         "column_titles": [str] (optional), "note": str (optional)}"""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(LETTER),
        leftMargin=1.4*cm, rightMargin=1.4*cm,
        topMargin=1.4*cm, bottomMargin=1.4*cm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#0f172a"))
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748b"))
    heading_style = ParagraphStyle(
        "heading", parent=styles["Heading2"], fontSize=12,
        textColor=colors.HexColor("#1e293b"), spaceBefore=12, spaceAfter=6)
    note_style = ParagraphStyle(
        "note", parent=styles["Normal"], fontSize=8,
        textColor=colors.HexColor("#94a3b8"), spaceAfter=6, italic=True)

    story: list = [
        Paragraph(f"<b>MediAd View</b> · {title}", title_style),
        Paragraph(subtitle, sub_style),
        Spacer(1, 6),
    ]

    header_bg = colors.HexColor("#1e293b")
    row_alt   = colors.HexColor("#f8fafc")

    for sec in sections:
        story.append(Paragraph(sec.get("heading", ""), heading_style))
        if sec.get("note"):
            story.append(Paragraph(sec["note"], note_style))
        columns = sec["columns"]
        titles = sec.get("column_titles") or columns
        rows = sec.get("rows", [])
        if not rows:
            story.append(Paragraph("<i>No data.</i>", note_style))
            continue
        # Truncate to 200 rows in PDF (keep it readable; XLSX/CSV have full data)
        display_rows = rows[:200]
        data = [titles] + [
            [_stringify(r.get(c))[:80] for c in columns] for r in display_rows
        ]
        # Compute column widths (rough)
        total_w = landscape(LETTER)[0] - 3*cm
        col_widths = [total_w / len(columns)] * len(columns)
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, row_alt]),
            ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl)
        if len(rows) > 200:
            story.append(Paragraph(
                f"(Showing first 200 of {len(rows)} rows — export XLSX/CSV for the full set.)",
                note_style))
        story.append(Spacer(1, 6))

    doc.build(story)
    return buf.getvalue()
