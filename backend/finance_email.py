"""
MediAd View — Finance Extensions: Email/SMTP, Settings, Users, Exports, Signatures, AR
"""
from fastapi import APIRouter, Depends, HTTPException, Body, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from datetime import datetime, date
from calendar import monthrange
from io import BytesIO
import os
import uuid
import asyncio
import base64
import hashlib
import csv
import smtplib
from email.message import EmailMessage
from email.utils import formataddr

import aiosmtplib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from cryptography.fernet import Fernet

from finance_pdf import generate_invoice_pdf, generate_deposit_pdf, generate_contract_pdf, COMPANY

ext_router = APIRouter(prefix="/api/finance")

# ============ ENCRYPTION (for SMTP password) ============
def _get_fernet():
    key = os.environ.get("FERNET_KEY")
    if not key:
        # Derive from JWT secret for stability across restarts
        secret = os.environ.get("JWT_SECRET", "fallback-secret-key-mediadview")
        key = base64.urlsafe_b64encode(hashlib.sha256(secret.encode()).digest())
    elif isinstance(key, str):
        key = key.encode()
    return Fernet(key)

def encrypt_password(plain: str) -> str:
    if not plain:
        return ""
    return _get_fernet().encrypt(plain.encode()).decode()

def decrypt_password(token: str) -> str:
    if not token:
        return ""
    try:
        return _get_fernet().decrypt(token.encode()).decode()
    except Exception:
        return ""


# ============ MODELS ============
class SmtpSettings(BaseModel):
    smtp_host: str = "smtp.titan.email"
    smtp_port: int = 587
    smtp_user: str = ""
    smtp_password: str = ""
    smtp_use_tls: bool = True
    from_name: str = "MediAd View Billing"
    from_email: str = ""
    reply_to: Optional[str] = ""
    enabled: bool = False

class SendInvoiceRequest(BaseModel):
    to: Optional[str] = None  # override recipient
    cc: Optional[str] = None
    custom_message: Optional[str] = ""

class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    role: str  # superadmin, admin, accounting, sales, technical, viewer
    phone: Optional[str] = ""

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    active: Optional[bool] = None
    password: Optional[str] = None


# ============ EMAIL HTML TEMPLATE ============
def fmt_money(v):
    return "$" + f"{float(v or 0):,.2f}"

def fmt_date(s):
    if not s: return ""
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%B %d, %Y")
    except Exception:
        return s

def render_invoice_email_html(inv: dict, client: dict, base_url: str = "") -> str:
    """Intermedia-style branded invoice email."""
    period = ""
    if inv.get("period_start") and inv.get("period_end"):
        period = f"{fmt_date(inv['period_start'])} – {fmt_date(inv['period_end'])}"
    return f"""<!DOCTYPE html>
<html><body style="margin:0;padding:0;background:#f1f5f9;font-family:'Helvetica','Arial',sans-serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f1f5f9;padding:24px 12px">
  <tr><td align="center">
    <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 4px 16px rgba(15,23,42,.08);max-width:600px">
      <!-- Header -->
      <tr><td style="background:linear-gradient(135deg,#2563eb 0%,#1e40af 100%);padding:32px 32px 24px;text-align:center;color:#fff">
        <div style="font-size:11px;font-weight:700;letter-spacing:2.5px;opacity:.85;margin-bottom:10px">MEDIAD VIEW · ADVERTISING SOLUTION</div>
        <div style="font-size:22px;font-weight:700;color:#fff;margin-bottom:6px">Your Monthly Invoice</div>
        <div style="font-size:13px;opacity:.9">Invoice #{inv.get('invoice_number','')} · Period {period}</div>
      </td></tr>

      <!-- Greeting -->
      <tr><td style="padding:32px 32px 0">
        <div style="font-size:14px;color:#0f172a;margin-bottom:18px">Hello <strong>{client.get('representative','—')}</strong>,</div>
        <div style="font-size:13.5px;color:#334155;line-height:1.6">Thank you for your continued business with <strong>MediAd View</strong>. Below is your invoice summary for this billing period. The detailed PDF is attached to this email.</div>
      </td></tr>

      <!-- Summary box -->
      <tr><td style="padding:24px 32px 0">
        <table width="100%" cellpadding="0" cellspacing="0" style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px">
          <tr><td style="padding:18px 22px">
            <div style="font-size:10.5px;font-weight:700;color:#1e40af;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:14px">INVOICE SUMMARY</div>
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr><td style="padding:5px 0;font-size:13px;color:#475569">Client</td><td align="right" style="padding:5px 0;font-size:13px;font-weight:600;color:#0f172a">{client.get('business_name','—')}</td></tr>
              <tr><td style="padding:5px 0;font-size:13px;color:#475569">Invoice #</td><td align="right" style="padding:5px 0;font-size:13px;font-weight:600;color:#0f172a">{inv.get('invoice_number','')}</td></tr>
              <tr><td style="padding:5px 0;font-size:13px;color:#475569">Issue Date</td><td align="right" style="padding:5px 0;font-size:13px;font-weight:600;color:#0f172a">{fmt_date(inv.get('issue_date',''))}</td></tr>
              <tr><td style="padding:5px 0;font-size:13px;color:#475569">Due Date</td><td align="right" style="padding:5px 0;font-size:13px;font-weight:600;color:#0f172a">{fmt_date(inv.get('due_date',''))}</td></tr>
              <tr><td colspan="2" style="padding:10px 0 0"><div style="border-top:1px solid #bfdbfe"></div></td></tr>
              <tr><td style="padding:10px 0 0;font-size:15px;font-weight:700;color:#0f172a">Amount Due</td><td align="right" style="padding:10px 0 0;font-size:22px;font-weight:800;color:#1e40af">{fmt_money(inv.get('balance', inv.get('total',0)))}</td></tr>
            </table>
          </td></tr>
        </table>
      </td></tr>

      <!-- CTA Button -->
      <tr><td style="padding:24px 32px;text-align:center">
        <a href="{base_url}/api/finance/invoices/{inv.get('id','')}/render" style="display:inline-block;background:#2563eb;color:#fff;padding:13px 32px;border-radius:8px;text-decoration:none;font-weight:700;font-size:14px;box-shadow:0 4px 12px rgba(37,99,235,.3)">View Invoice Online</a>
      </td></tr>

      <!-- Payment info -->
      <tr><td style="padding:0 32px 24px">
        <table width="100%" cellpadding="0" cellspacing="0" style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px">
          <tr><td style="padding:16px 20px">
            <div style="font-size:11px;font-weight:700;color:#1e40af;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px">Payment Methods</div>
            <div style="font-size:12.5px;color:#334155;line-height:1.7">
              <strong>Bank:</strong> {COMPANY['bank_name']}<br>
              <strong>Account #:</strong> {COMPANY['account_number']}<br>
              <strong>Routing #:</strong> {COMPANY['routing']}<br>
              Make payable to: <strong>{COMPANY['name']}</strong>
            </div>
          </td></tr>
        </table>
      </td></tr>

      <!-- Inquiries -->
      <tr><td style="padding:0 32px 28px">
        <div style="font-size:12px;color:#64748b;line-height:1.6;text-align:center">
          <strong style="color:#0f172a">Billing Inquiries</strong><br>
          Phone: <strong>{COMPANY['phone_1']}</strong> · {COMPANY['phone_2']}<br>
          <a href="https://{COMPANY['website']}" style="color:#2563eb;text-decoration:none">{COMPANY['website']}</a>
        </div>
      </td></tr>

      <!-- Footer -->
      <tr><td style="background:#0f172a;padding:18px 32px;text-align:center">
        <div style="font-size:11px;color:#94a3b8;line-height:1.5">
          © {datetime.utcnow().year} {COMPANY['name']} · {COMPANY['address_line1']}, {COMPANY['address_line2']}
        </div>
      </td></tr>
    </table>
  </td></tr>
</table>
</body></html>"""


# ============ ROUTES FACTORY ============
def create_finance_extensions(db, get_current_user):
    async def require_finance(user: dict = Depends(get_current_user)):
        if user.get("role") not in ("superadmin", "admin", "accounting", "sales", "viewer", "technical"):
            raise HTTPException(403, "Finance access required")
        return user

    async def require_admin(user: dict = Depends(get_current_user)):
        if user.get("role") not in ("superadmin", "admin"):
            raise HTTPException(403, "Admin access required")
        return user

    # ============ SMTP SETTINGS ============
    @ext_router.get("/settings/email")
    async def get_email_settings(user: dict = Depends(require_admin)):
        s = await db.fin_settings.find_one({"_id": "email"})
        if not s:
            return SmtpSettings(smtp_user="", from_email="").dict()
        s.pop("_id", None)
        # Don't return the encrypted password
        if s.get("smtp_password"):
            s["smtp_password"] = "********"  # placeholder showing it's set
            s["password_set"] = True
        else:
            s["password_set"] = False
        return s

    @ext_router.put("/settings/email")
    async def update_email_settings(data: SmtpSettings, user: dict = Depends(require_admin)):
        doc = data.dict()
        # Encrypt password if provided (and not placeholder)
        if doc.get("smtp_password") and doc["smtp_password"] != "********":
            doc["smtp_password"] = encrypt_password(doc["smtp_password"])
        else:
            # Keep existing
            existing = await db.fin_settings.find_one({"_id": "email"})
            if existing and existing.get("smtp_password"):
                doc["smtp_password"] = existing["smtp_password"]
            else:
                doc["smtp_password"] = ""
        doc["updated_at"] = datetime.utcnow().isoformat()
        doc["updated_by"] = user.get("email", "")
        await db.fin_settings.update_one({"_id": "email"}, {"$set": doc}, upsert=True)
        return {"ok": True}

    @ext_router.post("/settings/email/test")
    async def test_email(payload: dict = Body(...), user: dict = Depends(require_admin)):
        to_addr = payload.get("to") or user.get("email")
        s = await db.fin_settings.find_one({"_id": "email"})
        if not s:
            raise HTTPException(400, "Email settings not configured")
        try:
            msg = EmailMessage()
            msg["From"] = formataddr((s.get("from_name", "MediAd View"), s.get("from_email") or s.get("smtp_user")))
            msg["To"] = to_addr
            msg["Subject"] = "MediAd View — SMTP Test Email"
            msg.set_content("This is a test email from your MediAd View system. SMTP is configured correctly!")
            msg.add_alternative(
                '<div style="font-family:Arial;padding:20px"><h2 style="color:#2563eb">✓ SMTP Test Successful</h2>'
                '<p>Your MediAd View email integration is working correctly.</p>'
                '<p style="font-size:12px;color:#64748b">Sent from your billing system.</p></div>',
                subtype="html",
            )
            await _send_email_async(s, msg, to_addr)
            return {"ok": True, "message": f"Test email sent to {to_addr}"}
        except Exception as e:
            raise HTTPException(500, f"SMTP error: {str(e)}")

    async def _send_email_async(s, msg, to_addr):
        pwd = decrypt_password(s.get("smtp_password", ""))
        port = int(s.get("smtp_port", 587))
        use_tls = port == 465
        start_tls = not use_tls
        await aiosmtplib.send(
            msg,
            hostname=s.get("smtp_host", "smtp.titan.email"),
            port=port,
            username=s.get("smtp_user"),
            password=pwd,
            use_tls=use_tls,
            start_tls=start_tls,
            timeout=30,
        )

    # ============ PDF GENERATION ============
    @ext_router.get("/invoices/{invoice_id}/pdf")
    async def invoice_pdf(invoice_id: str, user: dict = Depends(require_finance)):
        inv = await db.fin_invoices.find_one({"id": invoice_id})
        if not inv:
            raise HTTPException(404, "Invoice not found")
        client = await db.fin_clients.find_one({"id": inv["client_id"]}) or {}
        pdf_bytes = generate_invoice_pdf(inv, client)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="Invoice_{inv.get("invoice_number","")}.pdf"'},
        )

    @ext_router.get("/contracts/{contract_id}/pdf")
    async def contract_pdf(contract_id: str, user: dict = Depends(require_finance)):
        ct = await db.fin_contracts.find_one({"id": contract_id})
        if not ct:
            raise HTTPException(404, "Contract not found")
        client = await db.fin_clients.find_one({"id": ct["client_id"]}) or {}
        pdf_bytes = generate_contract_pdf(ct, client)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="Contract_{ct.get("contract_number","")}.pdf"'},
        )

    @ext_router.get("/deposits/{deposit_id}/pdf")
    async def deposit_pdf(deposit_id: str, user: dict = Depends(require_finance)):
        dep = await db.fin_deposits.find_one({"id": deposit_id})
        if not dep:
            raise HTTPException(404, "Deposit not found")
        client = await db.fin_clients.find_one({"id": dep["client_id"]}) or {}
        pdf_bytes = generate_deposit_pdf(dep, client)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="Deposit_{dep.get("receipt_number","")}.pdf"'},
        )

    # ============ SEND INVOICE BY EMAIL ============
    @ext_router.post("/invoices/{invoice_id}/send")
    async def send_invoice_email(invoice_id: str, payload: SendInvoiceRequest = Body(...),
                                  user: dict = Depends(require_admin)):
        inv = await db.fin_invoices.find_one({"id": invoice_id})
        if not inv:
            raise HTTPException(404, "Invoice not found")
        client = await db.fin_clients.find_one({"id": inv["client_id"]}) or {}

        to_addr = (payload.to or "").strip() or client.get("email", "")
        if not to_addr:
            raise HTTPException(400, "Recipient email not provided. Please add one to the client profile or pass 'to' in request.")

        s = await db.fin_settings.find_one({"_id": "email"})
        if not s or not s.get("enabled"):
            raise HTTPException(400, "Email is not enabled. Configure SMTP in Settings → Email first.")

        # Build base URL for "View Online" link
        base_url = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")

        # Generate PDF attachment
        pdf_bytes = generate_invoice_pdf(inv, client)
        pdf_filename = f"Invoice_{inv.get('invoice_number','')}.pdf"

        # Build message
        msg = EmailMessage()
        msg["From"] = formataddr((s.get("from_name", "MediAd View Billing"), s.get("from_email") or s.get("smtp_user")))
        msg["To"] = to_addr
        if payload.cc:
            msg["Cc"] = payload.cc
        if s.get("reply_to"):
            msg["Reply-To"] = s["reply_to"]
        msg["Subject"] = f"Your MediAd View Invoice {inv.get('invoice_number','')} — {fmt_money(inv.get('balance', inv.get('total',0)))} due"

        # Plain text fallback
        text_body = (
            f"Hello {client.get('representative','')},\n\n"
            f"Your invoice {inv.get('invoice_number','')} is attached.\n"
            f"Amount due: {fmt_money(inv.get('balance', inv.get('total',0)))}\n"
            f"Due date: {fmt_date(inv.get('due_date',''))}\n\n"
            f"Payment to:\n"
            f"  Bank: {COMPANY['bank_name']}\n"
            f"  Account #: {COMPANY['account_number']}\n"
            f"  Routing: {COMPANY['routing']}\n\n"
            f"Questions? {COMPANY['phone_1']}\n"
            f"{COMPANY['name']}\n"
        )
        if payload.custom_message:
            text_body = payload.custom_message + "\n\n" + text_body
        msg.set_content(text_body)

        html_body = render_invoice_email_html(inv, client, base_url=base_url)
        if payload.custom_message:
            html_body = html_body.replace(
                '<div style="font-size:13.5px;color:#334155;line-height:1.6">',
                f'<div style="font-size:13.5px;color:#334155;line-height:1.6;padding:14px;background:#fef9c3;border-left:3px solid #f59e0b;border-radius:6px;margin-bottom:14px">{payload.custom_message}</div><div style="font-size:13.5px;color:#334155;line-height:1.6">',
                1,
            )
        msg.add_alternative(html_body, subtype="html")
        msg.get_payload()[1].add_related  # noqa
        msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf", filename=pdf_filename)

        try:
            await _send_email_async(s, msg, to_addr)
        except Exception as e:
            raise HTTPException(500, f"Failed to send: {str(e)}")

        # Log the send
        await db.fin_invoices.update_one(
            {"id": invoice_id},
            {"$set": {"last_sent_at": datetime.utcnow().isoformat(),
                      "last_sent_to": to_addr,
                      "sent_count": inv.get("sent_count", 0) + 1}},
        )
        return {"ok": True, "sent_to": to_addr}

    # ============ ACCOUNTS RECEIVABLE ============
    @ext_router.get("/accounts-receivable")
    async def accounts_receivable(user: dict = Depends(require_finance)):
        today = datetime.utcnow().date().isoformat()
        # Mark overdue
        await db.fin_invoices.update_many(
            {"status": "pending", "due_date": {"$lt": today}},
            {"$set": {"status": "overdue"}},
        )
        # Get all unpaid invoices
        invs = await db.fin_invoices.find({"status": {"$in": ["pending", "overdue"]}}).sort("due_date", 1).to_list(2000)
        client_ids = list({i["client_id"] for i in invs})
        clients = {c["id"]: c for c in await db.fin_clients.find({"id": {"$in": client_ids}}).to_list(2000)}
        # Group by client
        by_client = {}
        for i in invs:
            i.pop("_id", None)
            cid = i["client_id"]
            cl = clients.get(cid, {})
            if cid not in by_client:
                by_client[cid] = {
                    "client_id": cid,
                    "client_name": cl.get("business_name", "—"),
                    "representative": cl.get("representative", ""),
                    "email": cl.get("email", ""),
                    "phone": cl.get("phone", ""),
                    "invoices": [],
                    "total_due": 0,
                    "overdue_count": 0,
                    "oldest_due": None,
                }
            balance = i.get("balance", i.get("total", 0))
            by_client[cid]["invoices"].append(i)
            by_client[cid]["total_due"] += balance
            if i["status"] == "overdue":
                by_client[cid]["overdue_count"] += 1
            if not by_client[cid]["oldest_due"] or i.get("due_date", "") < by_client[cid]["oldest_due"]:
                by_client[cid]["oldest_due"] = i.get("due_date")
        result = sorted(by_client.values(), key=lambda x: x["total_due"], reverse=True)
        summary = {
            "total_clients_owing": len(result),
            "total_ar": sum(c["total_due"] for c in result),
            "total_overdue_invoices": sum(c["overdue_count"] for c in result),
            "total_open_invoices": sum(len(c["invoices"]) for c in result),
        }
        return {"summary": summary, "clients": result}

    # ============ CONTRACT SIGNATURE ============
    @ext_router.post("/contracts/{contract_id}/sign")
    async def sign_contract(contract_id: str, payload: dict = Body(...),
                            user: dict = Depends(require_finance)):
        """Save signature(s) on a contract. Payload: {lessor_signature, lessee_signature} (data URLs)."""
        upd = {}
        if "lessor_signature" in payload:
            upd["lessor_signature"] = payload["lessor_signature"]
        if "lessee_signature" in payload:
            upd["lessee_signature"] = payload["lessee_signature"]
        if upd:
            upd["signed_at"] = datetime.utcnow().strftime("%B %d, %Y")
            await db.fin_contracts.update_one({"id": contract_id}, {"$set": upd})
        return {"ok": True, "signed_at": upd.get("signed_at")}

    # ============ USER MANAGEMENT ============
    @ext_router.get("/users")
    async def list_users(user: dict = Depends(require_admin)):
        items = await db.users.find({}, {"password_hash": 0}).sort("created_at", -1).to_list(500)
        for u in items:
            u.pop("_id", None)
        return items

    @ext_router.post("/users")
    async def create_user(data: UserCreate, user: dict = Depends(require_admin)):
        # Verify role allowed
        allowed_roles = {"admin", "accounting", "sales", "technical", "viewer"}
        if user.get("role") == "superadmin":
            allowed_roles.add("superadmin")
        if data.role not in allowed_roles:
            raise HTTPException(400, f"Invalid role. Allowed: {sorted(allowed_roles)}")
        # Check email unique
        if await db.users.find_one({"email": data.email}):
            raise HTTPException(400, "Email already registered")
        import bcrypt
        password_hash = bcrypt.hashpw(data.password.encode(), bcrypt.gensalt()).decode()
        doc = {
            "id": str(uuid.uuid4()),
            "name": data.name,
            "email": data.email,
            "phone": data.phone or "",
            "password_hash": password_hash,
            "role": data.role,
            "active": True,
            "created_at": datetime.utcnow().isoformat(),
            "created_by": user.get("email", ""),
        }
        await db.users.insert_one(doc)
        doc.pop("password_hash", None)
        doc.pop("_id", None)
        return doc

    @ext_router.put("/users/{user_id}")
    async def update_user(user_id: str, data: UserUpdate, user: dict = Depends(require_admin)):
        upd = {k: v for k, v in data.dict().items() if v is not None and k != "password"}
        if data.password:
            import bcrypt
            upd["password_hash"] = bcrypt.hashpw(data.password.encode(), bcrypt.gensalt()).decode()
        if upd:
            upd["updated_at"] = datetime.utcnow().isoformat()
            await db.users.update_one({"id": user_id}, {"$set": upd})
        return {"ok": True}

    @ext_router.delete("/users/{user_id}")
    async def delete_user(user_id: str, user: dict = Depends(require_admin)):
        u = await db.users.find_one({"id": user_id})
        if not u:
            raise HTTPException(404, "User not found")
        if u.get("email") == user.get("email"):
            raise HTTPException(400, "Cannot delete yourself")
        await db.users.update_one({"id": user_id}, {"$set": {"active": False}})
        return {"ok": True}

    # ============ EXPORTS (Excel / CSV) ============
    def _xlsx_response(rows: List[List], headers: List[str], sheet_name: str, filename: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="0F172A")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, 1):
            maxw = max([len(str(h))] + [len(str(r[i-1])) if i-1 < len(r) and r[i-1] is not None else 0 for r in rows])
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(maxw + 3, 12), 60)
        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'})

    @ext_router.get("/export/invoices.xlsx")
    async def export_invoices(user: dict = Depends(require_finance)):
        items = await db.fin_invoices.find().sort("issue_date", -1).to_list(5000)
        client_map = {c["id"]: c for c in await db.fin_clients.find().to_list(5000)}
        rows = []
        for i in items:
            cl = client_map.get(i.get("client_id"), {})
            rows.append([
                i.get("invoice_number",""), cl.get("business_name",""),
                cl.get("representative",""), cl.get("email",""),
                i.get("issue_date",""), i.get("due_date",""),
                i.get("period_start",""), i.get("period_end",""),
                float(i.get("subtotal",0)), float(i.get("tax",0)),
                float(i.get("total",0)), float(i.get("amount_paid",0)),
                float(i.get("balance",0)), i.get("status",""),
            ])
        return _xlsx_response(rows, [
            "Invoice #","Client","Representative","Email","Issue Date","Due Date",
            "Period Start","Period End","Subtotal","Tax","Total","Amount Paid","Balance","Status"
        ], "Invoices", f"Invoices_{datetime.utcnow().strftime('%Y%m%d')}")

    @ext_router.get("/export/payments.xlsx")
    async def export_payments(user: dict = Depends(require_finance)):
        items = await db.fin_payments.find().sort("date", -1).to_list(5000)
        cm = {c["id"]: c.get("business_name","") for c in await db.fin_clients.find().to_list(5000)}
        rows = [[i.get("date",""), cm.get(i.get("client_id"),""), i.get("method",""), i.get("reference",""),
                 float(i.get("amount",0)), i.get("notes","")] for i in items]
        return _xlsx_response(rows, ["Date","Client","Method","Reference","Amount","Notes"], "Payments",
                              f"Payments_{datetime.utcnow().strftime('%Y%m%d')}")

    @ext_router.get("/export/expenses.xlsx")
    async def export_expenses(user: dict = Depends(require_finance)):
        items = await db.fin_expenses.find().sort("date", -1).to_list(5000)
        rows = [[i.get("date",""), i.get("category",""), i.get("description",""), i.get("vendor",""),
                 float(i.get("amount",0)), i.get("payment_method",""), i.get("notes","")] for i in items]
        return _xlsx_response(rows, ["Date","Category","Description","Vendor","Amount","Method","Notes"],
                              "Expenses", f"Expenses_{datetime.utcnow().strftime('%Y%m%d')}")

    @ext_router.get("/export/clients.xlsx")
    async def export_clients(user: dict = Depends(require_finance)):
        items = await db.fin_clients.find().sort("business_name", 1).to_list(5000)
        rows = []
        for c in items:
            total_screens = sum(sum(sc.get("units",1) for sc in (loc.get("screens",[]) or [])) for loc in (c.get("locations",[]) or []))
            rows.append([
                c.get("business_name",""), c.get("representative",""), c.get("email",""),
                c.get("phone",""), c.get("address_line1",""), c.get("city",""), c.get("state",""),
                c.get("zip",""), len(c.get("locations",[]) or []), total_screens, c.get("status",""),
                c.get("created_at","")[:10],
            ])
        return _xlsx_response(rows, [
            "Business","Representative","Email","Phone","Address","City","State","Zip",
            "Locations","Total Screens","Status","Created"
        ], "Clients", f"Clients_{datetime.utcnow().strftime('%Y%m%d')}")

    @ext_router.get("/export/accounts-receivable.xlsx")
    async def export_ar(user: dict = Depends(require_finance)):
        ar_data = await accounts_receivable(user)
        rows = []
        for c in ar_data["clients"]:
            for inv in c["invoices"]:
                rows.append([
                    c["client_name"], c["representative"], c["email"], c["phone"],
                    inv["invoice_number"], inv.get("issue_date",""), inv.get("due_date",""),
                    float(inv.get("total",0)), float(inv.get("balance", inv.get("total",0))),
                    inv["status"],
                ])
        return _xlsx_response(rows, [
            "Client","Representative","Email","Phone","Invoice #","Issue Date","Due Date",
            "Total","Balance Due","Status"
        ], "Accounts Receivable", f"AR_{datetime.utcnow().strftime('%Y%m%d')}")

    return ext_router
